import boto3
from time import strftime
import configparser
from botocore.exceptions import ClientError, WaiterError
from openpyxl import Workbook, load_workbook
import datetime, time

log_name = strftime('clean_log_' + "%Y-%b-%d_%H-%M-%S.log")
xlsx_name = None
console = False
Logfile = False

def get_config_regions():
    """
    read the region configuration from config.txt and validate it is correct
    :return: return list of region based on user config.txt
    """
    existing_regions = ('eu-north-1', 'ap-south-1', 'eu-west-3', 'eu-west-2', 'eu-west-1', 'ap-northeast-2',
                        'ap-northeast-1', 'sa-east-1', 'ca-central-1', 'ap-southeast-1', 'ap-southeast-2',
                        'eu-central-1', 'us-east-1', 'us-east-2', 'us-west-1', 'us-west-2')

    _log('INFO: Checking region config')
    config = configparser.ConfigParser()
    config.read('config.txt')

    if config['general'].getboolean('aws_regions_all'):
        _log('INFO: Regions from config file are - All regions')
        return existing_regions
    else:
        region_list = config['general']['aws_regions'].split(",")

        bad_region = [region for region in region_list if region.strip() not in existing_regions]
        if bad_region:
            _log(f"ERROR: Not found - {bad_region}, Please check your configuration")
        region_list = [region for region in region_list if region.strip() in existing_regions]
        _log(f"INFO: Valid regions from config file are - {region_list}")
        return region_list


def get_config(value, section,type='string'):
    '''
    General get config function, return
    :param type: whether to get the value as a bool or string from config.txt
    :return: requested value from config.xt
    '''
    config = configparser.ConfigParser()
    config.read('config.txt')
    if type == 'string':
        value = [config[section][value]]
    elif type == 'bool':
        value  = config[section].getboolean(value)

    _log(f"INFO: got {value} from section {section}")
    return value

def clean_ec2(regions, target_account ='Main', dry_run=True):
    _log(f"INFO: Starting EC2 cleaning for target {target_account}")

    cleanup_mode = get_config('Snapshots', 'cleanup')[0]
    _log(f'cleanup_mode={cleanup_mode}')

    # going over each region configured and checking for EC2
    for region in regions:
        _log(f"INFO: Checking EC2 instances in region - {region}")

        stop_list = []  # will store list of EC2 to be shutdown
        terminate_list = []  # will store list of EC2 to be terminated

        if target_account == 'Main':
            _log("INFO: Normal Client")
            ec2 = boto3.client('ec2', region_name=region.strip())
        else:
            _log("INFO: Assume Role Client")
            ec2 = _assume_role('ec2', region.strip(), 'client')

        response = ec2.describe_instances()
        ec2_instances = [i for instance in response['Reservations'] for i in
                         instance['Instances']]  # extract the list of instances from the response
        if not ec2_instances:
            _log(f'WARNING: region {region}: No EC2 instances found')
        else:
            _log(f'INFO: region {region}: Found EC2 instances')
            for instance in ec2_instances:

                instance['Account'] = target_account

                # update OperationDone based on cleanup mode selected
                operation = 'keep'
                if instance.get('Tags'):
                    Tags = {tag.get('Key'): tag.get('Value') for tag in instance.get('Tags')}

                    if 'keep' not in Tags:  # no keep tag
                        operation = 'Terminate'
                    elif Tags['keep'] != instance['InstanceId']:  # keep tag not equal snapID
                        operation = 'Terminate'
                    elif 'keep_state' not in Tags:# instance is tagged, check if need to shutdown
                        operation = 'Shutdown'
                else:
                    Tags = 'N/A'
                    operation = 'Terminate'

                # for keeptag_withdate option, if less the time param, disable delete
                instance['ec2_time'] = 'N/A'
                if cleanup_mode == 'keeptag_withdate':
                    instance['ec2_time'] = calc_day_time_delta(instance['LaunchTime'], True)
                    cleanup_mode_time = int(get_config('time', 'cleanup')[0])
                    if instance['ec2_time'] <= cleanup_mode_time:
                        operation = 'keep'

                # get tags and check what operation need to be done
                # if not instance.get('Tags'):
                #     Tags = 'N/A'
                #     operation = 'Terminate'
                # else:
                #     Tags = {tag.get('Key'): tag.get('Value') for tag in instance.get('Tags')}
                #     if Tags.get('keep') == 'on':
                #         operation = 'DoNothing'
                #     elif Tags.get('keep') == 'off' or Tags.get('keep') == '':
                #         operation = 'Shutdown'
                #     else:
                #         operation = 'Terminate'

                _log(f"INFO: instance: {instance}")
                print_results_xlsx(data=instance, sheetname='EC2', Tags=str(Tags), OperationDone=operation)

                if operation == 'Shutdown':
                    stop_list.append(instance['InstanceId'])
                elif operation == 'Terminate':
                    terminate_list.append(instance['InstanceId'])

            if stop_list:  # stop the instances
                _log(f'INFO: Stopping in region{region}: {stop_list}')
                try:
                    response = ec2.stop_instances(InstanceIds=stop_list, DryRun=dry_run)
                    _log(f"INFO: Stopping instance response {response}")
                except ClientError as e:
                    if "Request would have succeeded, but DryRun flag is set" not in str(e):
                        _log(f"ERROR: {e}")
                        print_results_xlsx(data=str(stop_list), sheetname='EC2', OperationDone='ERROR-Shutdown',
                                           error=str(e))

            if terminate_list:  # terminate the instances
                _log(f'INFO: Terminating in region{region}: {terminate_list}')
                try:
                    response = ec2.terminate_instances(InstanceIds=terminate_list, DryRun=dry_run)
                    _log(f"INFO: terminate instance response {response}")
                except ClientError as e:  # probably some permission error
                    if "Request would have succeeded, but DryRun flag is set" not in str(e):
                        _log(f"ERROR: {e}")
                        print_results_xlsx(data=str(terminate_list), sheetname='EC2', OperationDone='ERROR-Terminate',
                                           error=str(e))
                else:  # if termination raised no error, check if it finished (as volume are depended on this)
                    try:
                        waiter = ec2.get_waiter('instance_terminated')
                        waiter.wait(InstanceIds=terminate_list, WaiterConfig={'Delay': 15, 'MaxAttempts': 12},
                                    DryRun=dry_run)
                    except WaiterError as e:
                        _log(f"ERROR: {e}")
                        print_results_xlsx(data=str(terminate_list), sheetname='EC2',
                                           OperationDone='ERROR-waitTerminate',
                                           error=str(e))

        _log(f"INFO: region end: {region}")
    _log("INFO: existing clean_ec2()")



def clean_snapshot(regions, target_account ='Main', dry_run=True):
    """
    check for snapshot in all regions and delete them all
    :param dry_run: for BOTO3 call
    """
    _log(f"INFO: entering clean_snapshot() for target {target_account}")
    if target_account == 'Main':
        account = get_config('aws_account','aws_details')
    else:
        account = get_config('aws_account','aws_details_2nd')

    cleanup_mode = get_config('Snapshots','cleanup')[0]
    _log(f'cleanup_mode={cleanup_mode}')

    for region in regions:

        _log(f'INFO: Cleaning snapshots for {region}')
        if target_account == 'Main':
            _log("INFO: Normal Client")
            ec2 = boto3.client('ec2', region_name=region.strip())
        else:
            _log("INFO: Assume Role Client")
            ec2 = _assume_role('ec2', region.strip(), 'client')

        response = ec2.describe_snapshots(OwnerIds=account)
        _log(f'describe_snapshots response {response}')

        for snap in response['Snapshots']:

            snap['Operation'] = 'keep'
            snap['Account'] = target_account

            if snap.get('Tags'):
                snap['Tags'] = {tag.get('Key'): tag.get('Value') for tag in snap['Tags']}
            else: snap['Tags'] = 'None'

            #update OperationDone based on cleanup mode selected
            if snap.get('Tags'):
                if 'keep' not in snap['Tags']: #no keep tag
                    snap['Operation'] = 'Delete'
                elif snap['Tags']['keep'] !=snap['SnapshotId']: # keep tag not equal snapID
                        snap['Operation'] = 'Delete'
            else: snap['Operation'] = 'Delete' # no tags

            # for keeptag_withdate option, if less the time param, disable delete
            snap['snap_time']='N/A'
            if cleanup_mode == 'keeptag_withdate':
                snap['snap_time']  = calc_day_time_delta(snap['StartTime'],True)
                cleanup_mode_time = int(get_config('time', 'cleanup')[0])
                if snap['snap_time'] <= cleanup_mode_time:
                    snap['Operation'] = 'keep'

            try:
                _log(f"INFO: Found {snap['SnapshotId']} for volume: {snap['VolumeId']}, size {snap['VolumeSize']} GB")
                if snap['Operation'] == 'Delete':
                    ec2.delete_snapshot(SnapshotId=snap['SnapshotId'], DryRun=dry_run)
            except ClientError as e:
                if "Request would have succeeded, but DryRun flag is set" not in str(e):
                    _log(f'ERROR: {e}')
                    print_results_xlsx(data=snap, sheetname='Snapshots', region=region, error=e)
                else:
                    print_results_xlsx(data=snap, sheetname='Snapshots', region=region)
            else:
                print_results_xlsx(data=snap, sheetname='Snapshots', region=region)
    _log("INFO: existing clean_snapshot()")


def clean_volumes(regions, target_account ='Main', dry_run=True):
    """
    check for volumes in all regions and delete all state=available volumes
    :param dry_run: for BOTO 3 call
    """
    _log(f"INFO: entering clean_volumes() for target {target_account}")
    cleanup_mode = get_config('Volumes', 'cleanup')[0]  # get cleanup mode
    _log(f'cleanup_mode={cleanup_mode}')

    for region in regions:
        _log(f'INFO: Cleaning available volumes for {region}')
        if target_account == 'Main':
            ec2 = boto3.client('ec2', region_name=region.strip())
        else:
            ec2 = _assume_role('ec2', region.strip(), 'client')

        response = ec2.describe_volumes()
        for volume in response['Volumes']:
            volume['Account'] = target_account

            Tags = volume.get('Tags')
            if Tags:
                Tags = {tag.get('Key'): tag.get('Value') for tag in Tags}
            else: Tag = 'None'
            _log(f"INFO: Found volume in {volume['AvailabilityZone']}: {volume['VolumeId']}({volume['State']},"
                f" {volume['Iops']} IOPS, {volume['VolumeType']}) with Tag: {Tags}")

            # update OperationDone based on cleanup mode selected
            State = 'Keep'
            volume['volume_time'] = 'N/A'
            if volume['State'] == 'available':
                if not Tags:  # no tags at all, delete
                    State = 'Terminate'
                elif 'keep' not in Tags:  # no keep tag, delete
                    State = 'Terminate'
                elif Tags['keep'] != volume['VolumeId']:
                    State = 'Terminate'

                # for keeptag_withdate option, if less the time param, disable delete

                if cleanup_mode == 'keeptag_withdate':
                    volume['volume_time'] = calc_day_time_delta(volume['CreateTime'], True)
                    cleanup_mode_time = int(get_config('time', 'cleanup')[0])
                    if volume['volume_time'] <= cleanup_mode_time:
                        State = 'Keep'
            try:
                if State == 'Terminate':
                    _log('INFO: Deleting Volume')
                    ec2.delete_volume(VolumeId=volume['VolumeId'], DryRun=dry_run)
            except ClientError as e:
                if "Request would have succeeded, but DryRun flag is set" not in str(e):
                    _log(f'ERROR: {e}')
                    print_results_xlsx(data=volume, sheetname='Volumes', Tags=Tags, OperationDone=State, error=e)
                else:
                    print_results_xlsx(data=volume, sheetname='Volumes', Tags=Tags, OperationDone=State)
            else:
                print_results_xlsx(data=volume, sheetname='Volumes', Tags=Tags, OperationDone=State)

    _log("INFO: existing clean_volumes()")


def clean_images(regions, target_account ='Main', dry_run=True):
    """
    check for AMI's in all regions and Deregister if there is no tag keep
    :param dry_run: for BOTO 3 call
    """

    _log(f"INFO: entering clean_images() for target {target_account}")
    if target_account == 'Main':
        account = get_config('aws_account', 'aws_details')
    else:
        account = get_config('aws_account', 'aws_details_2nd')
    _log(account)

    cleanup_mode = get_config('Images','cleanup')[0] #get cleanup mode
    _log(f'cleanup_mode={cleanup_mode}')

    for region in regions:
        _log(f'INFO: Cleaning available images for {region}')
        if target_account == 'Main':
            ec2 = boto3.client('ec2', region_name=region.strip())
        else:
            ec2 = _assume_role('ec2', region.strip(), 'client')

        images = ec2.describe_images(Owners=account)
        if not images['Images']:
            _log(f'WARNING: no images found for {region}')
        else:
            for img in images['Images']:

                Tags = img.get('Tags')
                if Tags: #get tags
                    Tags = img.get('Tags')
                    Tags = {tag.get('Key'): tag.get('Value') for tag in Tags}

                # update OperationDone based on cleanup mode selected
                OperationDone = 'Keep'
                if not Tags: # no tags at all, delete
                    OperationDone = 'Deregister'
                elif 'keep' not in Tags: # no keep tag, delete
                    OperationDone = 'Deregister'
                elif Tags['keep']!= img['ImageId']:
                    OperationDone = 'Deregister'

                # for keeptag_withdate option, if less the time param, disable delete
                img['amitime'] ='N/A'
                if cleanup_mode == 'keeptag_withdate':
                    DATETIME_FORMAT_YMD_HMS = "%Y-%m-%dT%H:%M:%S.%fZ"
                    img['amitime'] = datetime.datetime.strptime(img['CreationDate'], DATETIME_FORMAT_YMD_HMS)
                    img['amitime'] = calc_day_time_delta(img['amitime'], False)
                    cleanup_mode_time = int(get_config('time', 'cleanup')[0])
                    if img['amitime'] <= cleanup_mode_time:
                        OperationDone = 'Keep'

                img['Account'] = target_account
                try:
                    if OperationDone == "Deregister":
                        ec2.deregister_image(ImageId=img['ImageId'], DryRun=dry_run)

                except ClientError as e:
                    if "Request would have succeeded, but DryRun flag is set" not in str(e):
                        print_results_xlsx(data=img, sheetname='Images', region=region, OperationDone=OperationDone,
                                           Tags=Tags, error=e)
                    else:
                        print_results_xlsx(data=img, sheetname='Images', region=region, OperationDone=OperationDone,
                                           Tags=Tags)
                else:
                    print_results_xlsx(data=img, sheetname='Images', region=region, OperationDone=OperationDone,
                                       Tags=Tags)
    _log("INFO: existing clean_images()")


def clean_sg(regions, target_account ='Main', dry_run=True):
    """
    Check each region for security groups with boto3, delete SG that are unused & untagged
    :param dry_run: used for boto call, to avoid actually deleting anything
    :return: None
    """
    _log(f"INFO: Cleaning SG for target {target_account}")
    headers = ["Region", "OwnerId", "SG Name", "SG Id", "VpcId", "FromPort",
               "ToPort", "IpProtocol", "Source", "Instances", "Tags", "OperationDone"]

    for region in regions:  # iterate over the region list and get the SG's
        if target_account == 'Main':
            ec2 = boto3.client('ec2', region_name=region.strip())
        else:
            ec2 = _assume_role('ec2', region.strip(), 'client')

        response = ec2.describe_security_groups()

        _log(f"INFO: Checking SG in region - {region}")

        security_group_record = {'Region': region}  # dict for the SG, will be send later to the report

        for sg in response['SecurityGroups']:  # iterate over all the SG in the current region and add data to dict
            _log(f"INFO: Found security group")
            _log(f"INFO: {sg}")

            security_group_record['GroupName'] = sg['GroupName']
            security_group_record['VpcId'] = sg.get('VpcId')
            security_group_record['Account'] = target_account

            security_group_record['Instances'] = ''

            # get instances so we have SG -> relation
            instances_for_sg = ec2.describe_instances(
                Filters=[{'Name': 'instance.group-id', 'Values': [sg.get('GroupId'), ]}, ])
            instances_for_sg = [i for instance in instances_for_sg['Reservations'] for i in
                                instance['Instances']]
            instances_for_sg = [instance['InstanceId'] for instance in instances_for_sg]

            # set OperationDone to N/A, will be updated later if we delete
            security_group_record['OperationDone'] = 'N/A'
            security_group_record['GroupId'] = sg.get('GroupId')
            if sg.get('Tags'):
                security_group_record['Tags'] = {tag.get('Key'): tag.get('Value') for tag in sg.get('Tags')}
            else: security_group_record['Tags'] = 'None'

            delete_error = "None"
            if not instances_for_sg:  # if no instances found, check for tag and update 'OperationDone'
                security_group_record['Instances'] = 'N/A'

                sg_tag_no_delete = False

                if sg.get('Tags'):  # check if there are any tags at all
                    for tag in sg.get('Tags'):  # check for the relevant tag
                        if tag.get('Key') == 'keep':
                            _log('INFO: Found no delete tag(keep)')
                            sg_tag_no_delete = True  # don't delete

                if security_group_record['GroupName'] == 'default':  # cant delete default groups
                    sg_tag_no_delete = True  # don't delete

                if not sg_tag_no_delete:
                    security_group_record['OperationDone'] = 'Deleting'
                    _log(f'INFO: removing sg - {sg.get("GroupId")}')
                    try:
                        ec2.delete_security_group(GroupId=sg.get('GroupId'), DryRun=dry_run)
                    except ClientError as e:
                        if "Request would have succeeded, but DryRun flag is set" not in str(e):

                            delete_error = e
                        else:
                            delete_error = 'None'

            else:
                security_group_record['Instances'] = ', '.join(instances_for_sg)  # convert instance list to string

            print_results_xlsx(data=security_group_record, sheetname='SG',
                               OperationDone=security_group_record['OperationDone'], error=delete_error)
        _log("INFO: Region END")


def clean_rds_instances(regions, target_account ='Main', dry_run=True):
    '''
    Clean all untagged RDS instance in the target region
    '''

    _log(f"INFO: entering clean_rds_instances() for target {target_account}")
    cleanup_mode = get_config('RDS', 'cleanup')[0]
    _log(f'cleanup_mode={cleanup_mode}')

    for region in regions:
        _log(f'INFO: Cleaning available RDS for {region}')
        if target_account == 'Main':
            rds = boto3.client('rds', region_name=region.strip())
        else:
            rds = _assume_role('rds', region.strip(), 'client')

        databases = rds.describe_db_instances()
        for db in databases['DBInstances']:
            db['region'] = region.strip()
            db['Account'] = target_account


            if db['BackupRetentionPeriod'] == 0: # related to automatic backup for excel
                db['BackupRetentionPeriod'] = 'Disable'
            else:
                db['BackupRetentionPeriod'] = 'Enable'

            db['operation'] = 'DoNothing'
            if not db.get('TagList'): # no tags at all, terminate
                db['TagList'] = 'None'
                db['operation'] = 'Terminate'

            else: #tags exist
                db['TagList'] = {tag.get('Key'): tag.get('Value') for tag in db.get('TagList')}
                if db['TagList'].get('keep') != db['DBInstanceIdentifier']: #check if keep != db ID and if so delete
                    db['operation'] = 'Terminate'
                elif not db['TagList'].get('keep_state'): #db is tagged, check if need to shutdown
                    db['operation'] = 'Shutdown'

            #check if keeptag_withdate and update operaion if needed
            db['db_time'] = 'N/A'
            if cleanup_mode == 'keeptag_withdate':
                db['db_time'] = calc_day_time_delta(db['InstanceCreateTime'], True)
                cleanup_mode_time = int(get_config('time', 'cleanup')[0])
                if db['db_time'] <= cleanup_mode_time:
                    db['operation'] = 'Ignore'
            try:
                if not dry_run:
                    if db['operation']  == 'Terminate' :
                        rds.delete_db_instance(DBInstanceIdentifier=db['DBInstanceIdentifier'], SkipFinalSnapshot=True,
                                               DeleteAutomatedBackups=True)
                    elif db['operation']  == 'Shutdown':
                        rds.stop_db_instance(DBInstanceIdentifier=db['DBInstanceIdentifier'])

            except ClientError as e:
                db['error'] = e

            else:
                db['error'] = 'None'
            print_results_xlsx(data=db, sheetname='RDS Instances')


def clean_rds_instances_snaps(regions, target_account ='Main', dry_run=True):
    '''
    Clean all Manual RDS Snapshots from the target regions
    '''

    _log(f"INFO: entering clean_rds_instances_snaps() for target {target_account}")

    cleanup_mode = get_config('RDS_Snaps', 'cleanup')[0]
    _log(f'cleanup_mode={cleanup_mode}')

    for region in regions:
        _log(f'INFO: Cleaning available RDS snaps for {region}')
        if target_account == 'Main':
            rds = boto3.client('rds', region_name=region.strip())
        else:
            rds = _assume_role('rds', region.strip(), 'client')

        databases_snapshots = rds.describe_db_snapshots()

        for db_snap in databases_snapshots['DBSnapshots']:

            db_snap['region'] = region.strip()
            db_snap['Account'] = target_account

            if db_snap.get('TagList'): # check for tags
                db_snap['TagList'] = {tag.get('Key'): tag.get('Value') for tag in db_snap['TagList']}
            else: db_snap['TagList'] = 'None'
            db_snap['operation'] = 'Ignore'
            if db_snap['SnapshotType'] == 'manual': # can only delete manual snaps
                if 'keep' not in db_snap['TagList']: # if no keep tag delete
                    db_snap['operation'] = 'Delete'
                elif db_snap['TagList']['keep'] != db_snap['DBSnapshotIdentifier']: #if keep != snap-Id delete
                    db_snap['operation'] = 'Delete'

            # for keeptag_withdate option, if less the time param, disable delete
            db_snap['snap_time'] = 'N/A'
            if cleanup_mode == 'keeptag_withdate': # check for snapshot age
                db_snap['snap_time'] = calc_day_time_delta(db_snap['SnapshotCreateTime'], True)
                cleanup_mode_time = int(get_config('time', 'cleanup')[0])
                if db_snap['snap_time'] <= cleanup_mode_time:
                    db_snap['operation'] = 'Ignore'

            try:

                if not dry_run and db_snap['SnapshotType'] == 'manual' and db_snap['operation'] == 'Delete':
                    rds.delete_db_snapshot(DBSnapshotIdentifier=db_snap['DBSnapshotIdentifier'])

            except ClientError as e:
                db_snap['error'] = e

            else:
                db_snap['error'] = 'None'
            print_results_xlsx(data=db_snap, sheetname='RDS Snapshots')


def clean_S3_objects(target_account ='Main', dry_run=True):
    '''
    Clean all S3 objects from untagged bucket
    '''

    _log(f"INFO: entering clean_S3_objects() for target {target_account}")

    if target_account =='Main':
        s3 = boto3.client('s3')
    else:
        s3 = _assume_role('s3', 'us-east-1', 'client')

    bucket_list = s3.list_buckets()
    for bucket in bucket_list['Buckets']:
        bucket['failcount'] = 0
        bucket['Account'] = target_account

        if target_account == 'Main':
            s3cleanup = boto3.resource('s3')
        else:
            s3cleanup = _assume_role('s3', 'us-east-1', 'resource')

        _log(f" In bucket {bucket['Name']}, created boto3.resource, timestamp: {datetime.datetime.now()}")
        mybucket = s3cleanup.Bucket(bucket['Name'])
        keycount = sum(1 for _ in mybucket.objects.all())


        try:
            bucket['error']='None'
            tags = s3.get_bucket_tagging(Bucket=bucket['Name'])

        except ClientError as e:

            if 'NoSuchTagSet' in str(e) and keycount > 0:
                bucket['TagList'] = '[]'
                bucket['operation'] = 'Delete'

            elif keycount == 0:
                bucket['TagList'] = '[]'
                bucket['operation'] = 'DoNothing'

            else:
                bucket['error'] = e
                bucket['TagList'] = '[]'
                bucket['operation'] = 'N/A'

        else:
            tags['TagSet'] = {tag.get('Key'): tag.get('Value') for tag in tags['TagSet']}
            if 'keep' not in str(tags['TagSet']) and keycount > 0:
                bucket['TagList'] = tags['TagSet']
                bucket['operation'] = 'Delete'

            else:
                bucket['TagList'] = tags['TagSet']
                bucket['operation'] = 'DoNothing'




        if bucket['operation'] == 'Delete' and not dry_run:
            if keycount > 0:
                _log(f" keycount:  {keycount}, Doing delete(),  timestamp: {datetime.datetime.now()}")
                res = mybucket.objects.all().delete()
                if res[0].get('Errors'):
                    bucket['failcount'] += len(res[0]['Errors'])
                    bucket['error'] = res[0]['Errors'][0]['Code']
                _log(f" Finished delete(),  timestamp: {datetime.datetime.now()}")


        bucket['KeyCount'] = str(keycount)
        bucket['Account'] = target_account
        print_results_xlsx(data=bucket, sheetname='S3 Objects')


def create_xlsx(EC2=False, Volumes=False, Snapshots=False, Images=False, SG=False, RDS=False, RDS_Snaps=False, S3_Objects=False):
    '''
    Create the intial xlsx file with the releavant tabs based on the above param,
    '''
    _log('INFO: Creating excel')
    wb = Workbook()

    if EC2:
        # ws_ec2 = wb.active
        ws_ec2 = wb.create_sheet()
        ws_ec2.title = 'EC2'
        ws_ec2.append(
            ("OperationDone", "Age", "InstanceId", "InstanceType", "AvailabilityZone",
             "State", "Volumes",  "Account",
             "Tags"))

    if Volumes:
        ws_volumes = wb.create_sheet()
        ws_volumes.title = 'Volumes'
        ws_volumes.append(
            ("OperationDone","Age" ,"VolumeId", "AvailabilityZone", "State", "VolumeType", "Size(GB)", "Iops", "Account", "Tags",
             "Errors"))

    if Images:
        ws_images = wb.create_sheet()
        ws_images.title = 'Images'
        ws_images.append(
            ("OperationDone", "Age", "ImageId", "Name", "Region", "Account", "ImageType", "CreationDate", "Tags", "Errors"))

    if Snapshots:
        ws_snapshots = wb.create_sheet()
        ws_snapshots.title = 'Snapshots'
        ws_snapshots.append(("OperationDone","Age","SnapshotID", "VolumeId", "Region", "Account", "Tags", "Errors"))

    if SG:
        ws_images = wb.create_sheet()
        ws_images.title = 'SG'
        ws_images.append(
            ("OperationDone", "SG Id", "SG Name", "Account", "Region", "VpcId", "Instances", "Tags","Errors"))

    if RDS:
        ws_images = wb.create_sheet()
        ws_images.title = 'RDS Instances'
        ws_images.append(
            ("OperationDone", "Age","Region", "DBInstanceIdentifier", "DBInstanceStatus", "DBInstanceClass",
             "AllocatedStorage", "Automatic Backups", "Account",  "Tags", "Errors"))

    if RDS_Snaps:
        ws_images = wb.create_sheet()
        ws_images.title = 'RDS Snapshots'
        ws_images.append(
            ("OperationDone", "Age", "Region", "DBSnapshotIdentifier", "DBInstanceIdentifier", "SnapshotType", "Account",
             "Tags", "Errors"))

    if S3_Objects:
        ws_images = wb.create_sheet()
        ws_images.title = 'S3 Objects'
        ws_images.append(("OperationDone", "Bucket", "Tags", "Key Count", "Failed Count","Account", "Errors"))

    del wb['Sheet']
    wb.save(xlsx_name)
def print_results_xlsx(**kwargs):
    '''
    Add row to the xlsx file created with create_xlsx()
    :param kwargs: holds all info passed from the releavnt clean functions
    '''
    wb = load_workbook(xlsx_name)
    ws = wb[kwargs['sheetname']]

    error = kwargs.get('error')
    if kwargs['sheetname'] == 'Volumes':
        row = (
            kwargs['OperationDone'], kwargs['data']['volume_time'],  kwargs['data']['VolumeId'], kwargs['data']['AvailabilityZone'],
            kwargs['data']['State'], kwargs['data']['VolumeType'], kwargs['data']['Size'], kwargs['data']['Iops'], kwargs['data']['Account'],
            str(kwargs['Tags']), str(error)
        )
        ws.append(row)
        wb.save(xlsx_name)

    elif kwargs['sheetname'] == 'Snapshots':
        if not kwargs['data'].get('Tags'):
            kwargs['data']['Tags'] = 'N/A'
        row = (kwargs['data']['Operation'],kwargs['data']['snap_time'], kwargs['data']['SnapshotId'],kwargs['data']['VolumeId'], kwargs['region'], kwargs['data']['Account'], str(kwargs['data']['Tags']),
               str(error))
        ws.append(row)
        wb.save(xlsx_name)

    elif kwargs['sheetname'] == 'Images':
        row = (kwargs['OperationDone'], kwargs['data']["amitime"], kwargs['data']["ImageId"], kwargs['data']["Name"], kwargs['region'],
               kwargs['data']["Account"], kwargs['data']["ImageType"], kwargs['data']["CreationDate"],
               str(kwargs["Tags"]), str(error))
        ws.append(row)
        wb.save(xlsx_name)

    elif kwargs['sheetname'] == 'EC2' and error == None:

        volume_list = ''
        for volume in kwargs['data']['BlockDeviceMappings']:
            volume_list += f"{(volume['Ebs']['VolumeId'])}({volume['Ebs']['Status']}),  "

        # sg_list_name = ''
        # sg_list_id = ''
        # for sg in kwargs['data']['SecurityGroups']:
        #     sg_list_name += f"{sg['GroupName']},  "
        #     sg_list_id += f"{sg['GroupId']},  "

        row = (kwargs['OperationDone'], kwargs['data']['ec2_time'], kwargs['data']['InstanceId'], kwargs['data']['InstanceType'],
               kwargs['data']['Placement']['AvailabilityZone'],
               kwargs['data']['State']['Name'],
               volume_list, kwargs['data']['Account'],
               kwargs['Tags'])
        ws.append(row)
        wb.save(xlsx_name)
    elif kwargs['sheetname'] == 'EC2':
        ws.append((kwargs['OperationDone'], kwargs['data'], error))
        wb.save(xlsx_name)

    elif kwargs['sheetname'] == 'SG':

        row = (
            kwargs['OperationDone'], kwargs['data']["GroupId"], kwargs['data']["GroupName"], kwargs['data']["Account"],
            kwargs['data']['Region'], kwargs['data']["VpcId"], kwargs['data']["Instances"], str(kwargs['data']["Tags"]),
            str(error))

        ws.append(row)
        wb.save(xlsx_name)

    elif kwargs['sheetname'] == 'RDS Instances':

        row = (kwargs['data']["operation"], kwargs['data']["db_time"], kwargs['data']["region"], kwargs['data']["DBInstanceIdentifier"],
               kwargs['data']["DBInstanceStatus"], kwargs['data']["DBInstanceClass"],
               kwargs['data']["AllocatedStorage"], kwargs['data']["BackupRetentionPeriod"], kwargs['data']["Account"],
               str(kwargs['data']["TagList"]),
               str(kwargs['data']["error"]))

        ws.append(row)
        wb.save(xlsx_name)

    elif kwargs['sheetname'] == 'RDS Snapshots':
        # ("OperationDone", "Region", "DBSnapshotIdentifier", "DBInstanceIdentifier", "SnapshotType",
        #  "Tags", "Errors")
        row = ( kwargs['data']['operation'], kwargs['data']['snap_time'], kwargs['data']["region"], kwargs['data']["DBSnapshotIdentifier"], kwargs['data']["DBInstanceIdentifier"],
               kwargs['data']["SnapshotType"], kwargs['data']["Account"], str(kwargs['data']["TagList"]),
               str(kwargs['data']["error"]))

        ws.append(row)
        wb.save(xlsx_name)

    elif kwargs['sheetname'] == 'S3 Objects':
        # ("OperationDone", "Region", "DBSnapshotIdentifier", "DBInstanceIdentifier", "SnapshotType",
        #  "Tags", "Errors")
        row = ( kwargs['data']['operation'],  kwargs['data']["Name"], str(kwargs['data']["TagList"]), str(kwargs['data']["KeyCount"]),
                str(kwargs['data']["failcount"]),str(kwargs['data']["Account"]),str(kwargs['data']["error"]))

        ws.append(row)
        wb.save(xlsx_name)


def _log(line):
    '''
    Print to console and/or file
    '''

    if Logfile:
        with open(log_name, "a") as file:
            file.write(str(line) + '\n')
    if console:
        print(line)

def run_aws_cleanup(xlsxname, dry_run=True,EC2=False, Volumes=False, Snapshots=False, Images=False, SG=False, RDS=False, RDS_Snaps=False, S3_Objects=False, target_account='Main', createxlsx= True):
    '''
    Called from view.py, main function to run the cleanup operation, will call relevant cleanup function based on param recieved from views.py
    :param xlsxname: name of the xlsx file
    :param target_account: Main or Second
    :param createxlsx: does xlsx file need to be created or not
    :return:
    '''
    global xlsx_name
    global console
    global Logfile


    xlsx_name = xlsxname

    console = get_config('logs_console', 'general', 'bool')
    Logfile = get_config('logs_file', 'general', 'bool')
    _log(f"INFO: entering run_aws_cleanup() for target {target_account}")

    regions = get_config_regions()

    if createxlsx:
        create_xlsx(EC2=EC2, Volumes=Volumes, Snapshots=Snapshots, Images=Images, SG=SG, RDS=RDS, RDS_Snaps=RDS_Snaps, S3_Objects=S3_Objects)


    if EC2:
        clean_ec2(regions, target_account, dry_run)

    if Volumes:
        clean_volumes(regions, target_account, dry_run)

    if Images:
        clean_images(regions, target_account, dry_run)

    if Snapshots:
        clean_snapshot(regions, target_account, dry_run)

    if SG:
        clean_sg(regions, target_account, dry_run)
    if RDS:
        clean_rds_instances(regions, target_account, dry_run)

    if RDS_Snaps:
        clean_rds_instances_snaps(regions, target_account, dry_run)

    if S3_Objects:
        clean_S3_objects(target_account, dry_run)



def _assume_role(service,region='us-east-1', type='client'):
    '''
    Used to create assume role client for the requested service
    :param service: which AWS service
    :param type: default is client, can be resource(for S3)
    :return: the created client
    '''

    _log(f'INFO Entering _assume_role() for service {service} for region {region}')

    sts = boto3.client('sts')

    sts_response = sts.assume_role(
        RoleArn=get_config('role_to_assume', 'aws_details_2nd')[0].strip(),
        RoleSessionName='cleanersession' + service
    )

    new_session_id = sts_response["Credentials"]["AccessKeyId"]
    new_session_key = sts_response["Credentials"]["SecretAccessKey"]
    new_session_token = sts_response["Credentials"]["SessionToken"]

    if type=='client':
        assumed_client = boto3.client(
            service,
            region_name=region,
            aws_access_key_id=new_session_id,
            aws_secret_access_key=new_session_key,
            aws_session_token=new_session_token
        )
        return assumed_client
    else:
        assumed_resource = boto3.resource(
            service,
            region_name=region,
            aws_access_key_id=new_session_id,
            aws_secret_access_key=new_session_key,
            aws_session_token=new_session_token
        )
        return assumed_resource

def calc_day_time_delta(resourceTime,tz=True):
    '''
    calc the delta between current time and time param, used for datebased cleanup
    :param time: the time of the creation of the resource or snapshot
    :return: delta in days
    '''
    if tz:
        now = datetime.datetime.combine(datetime.datetime.now(datetime.timezone.utc) , datetime.time.max,tzinfo=datetime.timezone.utc)
        resourceTime =  datetime.datetime.combine(resourceTime,datetime.time.min, tzinfo=datetime.timezone.utc)
        return (now - resourceTime).days
    else: return (datetime.datetime.utcnow() -resourceTime).days


