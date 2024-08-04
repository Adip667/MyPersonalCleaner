# todo - connect to auzre
# todo - get list of RG
# todo - check tags
# todo - delete RG
import configparser
from time import strftime
from azure.mgmt.resource import ResourceManagementClient
from azure.identity import ClientSecretCredential
from openpyxl import Workbook, load_workbook

xlsx_name = None
console = False
Logfile = False
log_name = strftime('clean_log_Azure_' + "%Y-%b-%d_%H-%M-%S.log")

def create_xlsx():
    _log(f"INFO: entering create_xlsx()")
    try:
        wb = load_workbook(xlsx_name)

    except FileNotFoundError:
        wb = Workbook()

    ws_ec2 = wb.create_sheet()
    ws_ec2.title = 'azure RG'

    ws_ec2.append(
        ("OperationDone", "name", "location", "tags"))

    try:
        del wb['Sheet']
    except: pass
    wb.save(xlsx_name)


def print_results_xlsx(**kwargs):
    _log(f"INFO: entering print_results_xlsx()")
    wb = load_workbook(xlsx_name)
    ws = wb[kwargs['sheetname']]

    error = kwargs.get('error')

    row = (
        kwargs['OperationDone'], kwargs['Name'], kwargs['Location'], str(kwargs['Tags']))

    ws.append(row)
    wb.save(xlsx_name)


def get_config_account(value, section='azure_details',type='string'):
    _log(f"INFO: entering get_config_account()")
    config = configparser.ConfigParser()
    config.read('config.txt')
    if type == 'string':
        value = [config[section][value]]
    elif type == 'bool':
        value = config[section].getboolean(value)
    return value


def _log(line):
    '''
    Print to console and/or file
    '''

    if Logfile:
        with open(log_name, "a") as file:
            file.write(str(line) + '\n')
    if console:
        print(line)

def clean_az_rg(xlsxname, dry_run=True):

    client_secret = get_config_account('client_secret')[0]
    client_id = get_config_account('client_id')[0]
    tenant_id =get_config_account('tenant_id')[0]
    subscription_id = get_config_account('subscription_id')[0]

    credential = ClientSecretCredential(tenant_id=tenant_id, client_id=client_id, client_secret=client_secret)

    # subscription_client = SubscriptionClient(credential)
    # for sub in subscription_client.subscriptions.list():
    #     print(sub)
    #     print(sub.__getattribute__('subscription_id'))

    resource_client = ResourceManagementClient(credential, subscription_id)
    group_list = resource_client.resource_groups.list()

    global xlsx_name
    global console
    global Logfile


    console = get_config_account('logs_console', 'general','bool')
    Logfile = get_config_account('logs_file', 'general','bool')
    _log(f"INFO: entering clean_az_rg()")

    xlsx_name = xlsxname
    create_xlsx()
    _log(f'in clean_az_rg(), Dry_run is {dry_run}')
    for RG in group_list:

        if RG.__getattribute__('tags') is None :
            _log('Deleting Name: ' + RG.__getattribute__('name'))
            print_results_xlsx(OperationDone='Delete', Name=RG.__getattribute__('name'),
                               Location=RG.__getattribute__('location'), Tags=RG.__getattribute__('tags'), sheetname="azure RG")
            if not dry_run:

                delete_async_operation = resource_client.resource_groups.begin_delete(RG.__getattribute__('name'))

        elif 'keep' not in RG.__getattribute__('tags'):
            _log('Deleting Name: ' + RG.__getattribute__('name') + ', Tag ' + str(RG.__getattribute__('tags')))
            print_results_xlsx(OperationDone='Delete', Name=RG.__getattribute__('name'),
                               Location=RG.__getattribute__('location'), Tags=RG.__getattribute__('tags'), sheetname="azure RG")
            if not dry_run:

                delete_async_operation = resource_client.resource_groups.begin_delete((RG.__getattribute__('name')))

        else:
            _log('keeping: ' + RG.__getattribute__('name'))
            print_results_xlsx(OperationDone='Keep',Name=RG.__getattribute__('name'),Location=RG.__getattribute__('location'),Tags=RG.__getattribute__('tags'), sheetname="azure RG")

